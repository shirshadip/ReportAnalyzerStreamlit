import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


SUBJECTS = ["math", "physics", "cs", "english"]
SUBJECT_LABELS = {
    "math": "Mathematics",
    "physics": "Physics",
    "cs": "Computer Science",
    "english": "English",
}
MAX_MARKS_PER_SUBJECT = 100
TOTAL_MAX = MAX_MARKS_PER_SUBJECT * len(SUBJECTS)


class StudentAnalyzer:
    def __init__(self, df: pd.DataFrame):
        self.df = df.copy()
        self._enrich()

    def _enrich(self):
        df = self.df
        for s in SUBJECTS:
            df[s] = pd.to_numeric(df[s], errors="coerce").fillna(0).astype(int)
        df["total"]      = df[SUBJECTS].sum(axis=1)
        df["percentage"] = ((df["total"] / TOTAL_MAX) * 100).round(2)
        df["grade"]      = df["percentage"].apply(self._get_grade)
        df["rank"]       = df["total"].rank(method="min", ascending=False).astype(int)
        df.sort_values("rank", inplace=True)
        df.reset_index(drop=True, inplace=True)
        self.df = df

    @staticmethod
    def _get_grade(pct: float) -> str:
        if pct >= 90:   return "A+"
        elif pct >= 80: return "A"
        elif pct >= 70: return "B+"
        elif pct >= 60: return "B"
        elif pct >= 50: return "C"
        elif pct >= 40: return "D"
        else:           return "F"

    def full_analysis(self) -> dict:
        df = self.df

        ranked_students = df[[
            "id", "name", "math", "physics", "cs", "english",
            "total", "percentage", "grade", "rank"
        ]].to_dict(orient="records")

        topper_row = df[df["rank"] == 1].iloc[0]
        topper = {
            "name":       topper_row["name"],
            "total":      int(topper_row["total"]),
            "percentage": float(topper_row["percentage"]),
            "grade":      topper_row["grade"],
        }

        subject_averages = {
            SUBJECT_LABELS[s]: round(float(df[s].mean()), 2) for s in SUBJECTS
        }

        highest_idx = df["total"].idxmax()
        lowest_idx  = df["total"].idxmin()
        highest = {
            "name":       df.loc[highest_idx, "name"],
            "total":      int(df.loc[highest_idx, "total"]),
            "percentage": float(df.loc[highest_idx, "percentage"]),
        }
        lowest = {
            "name":       df.loc[lowest_idx, "name"],
            "total":      int(df.loc[lowest_idx, "total"]),
            "percentage": float(df.loc[lowest_idx, "percentage"]),
        }

        grade_dist = df["grade"].value_counts().to_dict()
        pass_count = int((df["percentage"] >= 40).sum())
        fail_count = int((df["percentage"] <  40).sum())

        return {
            "summary": {
                "total_students":           len(df),
                "class_average_percentage": round(float(df["percentage"].mean()), 2),
                "class_average_total":      round(float(df["total"].mean()), 2),
                "highest_percentage":       float(df["percentage"].max()),
                "lowest_percentage":        float(df["percentage"].min()),
                "pass_count":               pass_count,
                "fail_count":               fail_count,
            },
            "topper":          topper,
            "highest_scorer":  highest,
            "lowest_scorer":   lowest,
            "subject_averages": subject_averages,
            "grade_distribution": grade_dist,
            "ranked_students": ranked_students,
        }

    def _build_workbook(self) -> Workbook:
        analysis = self.full_analysis()
        wb = Workbook()

        header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        header_fill    = PatternFill("solid", fgColor="1A3C5E")
        accent_fill    = PatternFill("solid", fgColor="2196F3")
        highlight_fill = PatternFill("solid", fgColor="E3F2FD")
        gold_fill      = PatternFill("solid", fgColor="FFD700")
        white_fill     = PatternFill("solid", fgColor="FFFFFF")
        thin           = Side(style="thin", color="CCCCCC")
        thin_border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        center         = Alignment(horizontal="center", vertical="center")
        left           = Alignment(horizontal="left",   vertical="center")

        # ── Sheet 1: Summary ──
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.column_dimensions["A"].width = 32
        ws1.column_dimensions["B"].width = 22

        ws1.merge_cells("A1:B1")
        ws1["A1"] = "📊 Student Performance Report"
        ws1["A1"].font = Font(name="Calibri", bold=True, size=16, color="1A3C5E")
        ws1["A1"].alignment = center
        ws1["A1"].fill = PatternFill("solid", fgColor="E3F2FD")
        ws1.row_dimensions[1].height = 36

        ws1.merge_cells("A2:B2")
        ws1["A2"] = f"Total Students: {analysis['summary']['total_students']}"
        ws1["A2"].font = Font(name="Calibri", size=11, color="555555")
        ws1["A2"].alignment = center
        ws1.row_dimensions[2].height = 22

        metrics = [
            ("Class Average %",    f"{analysis['summary']['class_average_percentage']}%"),
            ("Class Average Total", f"{analysis['summary']['class_average_total']} / {TOTAL_MAX}"),
            ("Highest %",          f"{analysis['summary']['highest_percentage']}%"),
            ("Lowest %",           f"{analysis['summary']['lowest_percentage']}%"),
            ("Students Passed",     analysis["summary"]["pass_count"]),
            ("Students Failed",     analysis["summary"]["fail_count"]),
        ]

        row = 4
        ws1.merge_cells(f"A{row}:B{row}")
        ws1[f"A{row}"] = "CLASS OVERVIEW"
        ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        ws1[f"A{row}"].fill = accent_fill
        ws1[f"A{row}"].alignment = center
        ws1.row_dimensions[row].height = 24
        row += 1

        for i, (label, value) in enumerate(metrics):
            fill = highlight_fill if i % 2 == 0 else white_fill
            ws1[f"A{row}"] = label
            ws1[f"B{row}"] = value
            for col in ("A", "B"):
                ws1[f"{col}{row}"].fill = fill
                ws1[f"{col}{row}"].border = thin_border
            ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=11, color="1A3C5E")
            ws1[f"B{row}"].font = Font(name="Calibri", size=11)
            ws1[f"A{row}"].alignment = left
            ws1[f"B{row}"].alignment = center
            ws1.row_dimensions[row].height = 22
            row += 1

        row += 1
        ws1.merge_cells(f"A{row}:B{row}")
        ws1[f"A{row}"] = "🏆 CLASS TOPPER"
        ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="1A3C5E")
        ws1[f"A{row}"].fill = gold_fill
        ws1[f"A{row}"].alignment = center
        ws1.row_dimensions[row].height = 24
        row += 1

        cream = PatternFill("solid", fgColor="FFFDE7")
        for label, value in [
            ("Name",        analysis["topper"]["name"]),
            ("Total Score", f"{analysis['topper']['total']} / {TOTAL_MAX}"),
            ("Percentage",  f"{analysis['topper']['percentage']}%"),
            ("Grade",       analysis["topper"]["grade"]),
        ]:
            ws1[f"A{row}"] = label
            ws1[f"B{row}"] = value
            for col in ("A", "B"):
                ws1[f"{col}{row}"].fill = cream
                ws1[f"{col}{row}"].border = thin_border
            ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=11)
            ws1[f"B{row}"].font = Font(name="Calibri", size=11)
            ws1[f"A{row}"].alignment = left
            ws1[f"B{row}"].alignment = center
            row += 1

        row += 1
        ws1.merge_cells(f"A{row}:B{row}")
        ws1[f"A{row}"] = "📚 SUBJECT AVERAGES"
        ws1[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        ws1[f"A{row}"].fill = accent_fill
        ws1[f"A{row}"].alignment = center
        ws1.row_dimensions[row].height = 24
        row += 1

        for i, (subj, avg) in enumerate(analysis["subject_averages"].items()):
            fill = highlight_fill if i % 2 == 0 else white_fill
            ws1[f"A{row}"] = subj
            ws1[f"B{row}"] = avg
            for col in ("A", "B"):
                ws1[f"{col}{row}"].fill = fill
                ws1[f"{col}{row}"].border = thin_border
                ws1[f"{col}{row}"].font = Font(name="Calibri", size=11)
            ws1[f"A{row}"].alignment = left
            ws1[f"B{row}"].alignment = center
            row += 1

        # ── Sheet 2: Ranked Students ──
        ws2 = wb.create_sheet("Ranked Students")
        for i, w in enumerate([8, 22, 14, 14, 14, 14, 14, 14, 10, 8], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.merge_cells("A1:J1")
        ws2["A1"] = "STUDENT PERFORMANCE — RANKED LIST"
        ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        ws2["A1"].fill = header_fill
        ws2["A1"].alignment = center
        ws2.row_dimensions[1].height = 32

        for col, h in enumerate(
            ["Rank","Name","Math","Physics","CS","English","Total","Percentage","Grade","ID"], 1
        ):
            cell = ws2.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws2.row_dimensions[2].height = 24

        topper_fill = PatternFill("solid", fgColor="FFF9C4")
        for i, student in enumerate(analysis["ranked_students"]):
            row_num = i + 3
            is_topper = student["rank"] == 1
            row_fill  = topper_fill if is_topper else (highlight_fill if i % 2 == 0 else white_fill)
            for col, val in enumerate([
                student["rank"], student["name"],
                student["math"], student["physics"],
                student["cs"],   student["english"],
                student["total"], f"{student['percentage']}%",
                student["grade"], student["id"],
            ], 1):
                cell = ws2.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center
                cell.fill = row_fill
                cell.font = Font(name="Calibri", bold=is_topper, size=11)
                if col == 8:
                    pct = student["percentage"]
                    if pct >= 80:
                        cell.font = Font(name="Calibri", bold=True, color="1B5E20", size=11)
                    elif pct < 40:
                        cell.font = Font(name="Calibri", bold=True, color="B71C1C", size=11)
            ws2.row_dimensions[row_num].height = 20

        # ── Sheet 3: Grade Distribution ──
        ws3 = wb.create_sheet("Grade Distribution")
        for col in ("A", "B", "C"):
            ws3.column_dimensions[col].width = 16

        ws3.merge_cells("A1:C1")
        ws3["A1"] = "GRADE DISTRIBUTION"
        ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        ws3["A1"].fill = header_fill
        ws3["A1"].alignment = center
        ws3.row_dimensions[1].height = 32

        for col, h in enumerate(["Grade", "Count", "Percentage"], 1):
            cell = ws3.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        total_s    = analysis["summary"]["total_students"]
        grade_dist = analysis["grade_distribution"]
        for i, grade in enumerate(["A+", "A", "B+", "B", "C", "D", "F"]):
            count   = grade_dist.get(grade, 0)
            pct     = round((count / total_s * 100), 1) if total_s else 0
            row_num = i + 3
            fill    = highlight_fill if i % 2 == 0 else white_fill
            ws3.cell(row=row_num, column=1, value=grade)
            ws3.cell(row=row_num, column=2, value=count)
            ws3.cell(row=row_num, column=3, value=f"{pct}%")
            for col in range(1, 4):
                cell = ws3.cell(row=row_num, column=col)
                cell.fill = fill
                cell.alignment = center
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=11)

        chart = BarChart()
        chart.title        = "Grade Distribution"
        chart.y_axis.title = "Number of Students"
        chart.x_axis.title = "Grade"
        chart.add_data(Reference(ws3, min_col=2, min_row=2, max_row=9), titles_from_data=True)
        chart.set_categories(Reference(ws3, min_col=1, min_row=3, max_row=9))
        chart.shape = 4
        chart.width = 18
        chart.height = 12
        ws3.add_chart(chart, "E2")

        return wb

    def generate_excel_report_bytes(self) -> bytes:
        """Return Excel report as bytes — used by Streamlit download_button."""
        buffer = io.BytesIO()
        self._build_workbook().save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel_report(self, output_path: str = "/tmp/student_report.xlsx") -> str:
        """Save to disk and return path — kept for local/FastAPI use."""
        self._build_workbook().save(output_path)
        return output_path
